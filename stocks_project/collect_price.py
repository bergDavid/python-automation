from openpyxl import load_workbook
import os
import json
import requests

def apicollectprice(symbol):
    url = "https://yfapi.net/v7/finance/options/" + str(symbol)
    payload={}
    headers = {
    'X-API-KEY': 'OzSm0eWTFK37aCA2YaQRs2szoXcz0pcq90Vf4ivz',
    'Accept': 'application/json'
    }
    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        json_data = response.text
        json_load = json.loads(json_data)
        price = json_load["optionChain"]["result"][0]["quote"]["regularMarketPrice"]
        return price
    else:
        return "error"

#Update Stocks
excelpath = r"C:\Users\David\Desktop\my-stocks.xlsx"
wb = load_workbook(excelpath)
ws = wb.worksheets[0]
#stockprice = 10
for i in range(1, ws.max_row+1):
    stock = ws.cell(i,1).value
    stockprice = apicollectprice(stock)
    #stockprice += 10
    print(str(stock) + " : " + "{:.2f}".format(stockprice))
    ws.cell(i,2).value = stockprice

wb.save(excelpath)